"""Maintain application_tracker.csv (source of truth) and a pretty .xlsx.

- Reads dashboard.json (output of `python -m src daily-plan --export`)
- Merges fresh leads into application_tracker.csv
- Preserves Status / Applied Date / Notes from previous version
- Adds an "H1B Sponsor?" column based on the bundled
  Top100_H1B_Sponsor_Companies_with_Career_Links.csv
- Writes a styled application_tracker.xlsx (clickable apply links,
  Status dropdown, color-coded Status + H1B columns)

The CSV is the canonical version (easy to edit on GitHub web UI).
The xlsx is regenerated each run from the CSV + dashboard.
"""

from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

TRACKER_CSV = Path("application_tracker.csv")
TRACKER_XLSX = Path("application_tracker.xlsx")
H1B_CSV = Path("Top100_H1B_Sponsor_Companies_with_Career_Links.csv")

COLUMNS = [
    "Title",
    "Company",
    "Location",
    "H1B Sponsor?",
    "Apply Link",
    "Score",
    "Confidence",
    "Status",
    "Applied Date",
    "Notes",
    "First Seen",
]
STATUS_VALUES = ["New", "Applied", "Interview", "Rejected"]
STATUS_ORDER = {s: i for i, s in enumerate(STATUS_VALUES)}


def normalize(name: str) -> str:
    """Normalize a company name for fuzzy matching."""
    n = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode().lower()
    n = re.sub(r"[^a-z0-9]+", " ", n)
    for suffix in (
        "inc", "corp", "corporation", "llc", "ltd", "co", "company",
        "international", "group", "holdings", "technologies", "tech",
        "systems", "solutions", "services",
    ):
        n = re.sub(rf"\b{suffix}\b", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def load_h1b_sponsors() -> set[str]:
    """Extract sponsor company names from the bundled CSV.

    The file has a malformed format (everything on one line), so we extract
    company names by walking comma-separated tokens and treating any token
    that's NOT a URL/'None' as a company.
    """
    if not H1B_CSV.exists():
        return set()
    raw = H1B_CSV.read_text(encoding="utf-8", errors="ignore")
    sponsors: set[str] = set()
    # Tokenize by commas
    tokens = [t.strip() for t in raw.split(",")]
    for tok in tokens:
        if not tok:
            continue
        # A "company" token: doesn't start with http, isn't 'None', isn't header
        if tok.lower().startswith(("http://", "https://")):
            continue
        if tok in ("Company", "Career Page URL"):
            continue
        # Strip trailing 'None' (artifact in malformed CSV)
        if tok.endswith("None"):
            tok = tok[:-4].strip()
        if not tok:
            continue
        norm = normalize(tok)
        if norm:
            sponsors.add(norm)
    return sponsors


def is_h1b_sponsor(company: str, sponsors: set[str]) -> bool:
    if not company:
        return False
    norm = normalize(company)
    if not norm:
        return False
    if norm in sponsors:
        return True
    # Substring match (e.g., Serper says "Honeywell Aerospace" vs sponsor "Honeywell")
    for s in sponsors:
        if not s:
            continue
        if s in norm or norm in s:
            return True
    return False


def load_existing_tracker() -> dict[tuple[str, str], dict]:
    if not TRACKER_CSV.exists():
        return {}
    rows: dict[tuple[str, str], dict] = {}
    with TRACKER_CSV.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            key = (
                (row.get("Title") or "").strip().lower(),
                (row.get("Company") or "").strip().lower(),
            )
            rows[key] = row
    return rows


def write_csv(rows: list[dict]) -> None:
    with TRACKER_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for r in rows:
            writer.writerow({c: (r.get(c) or "") for c in COLUMNS})


def write_xlsx(rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("openpyxl unavailable; skipping xlsx generation", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    status_colors = {
        "New": "FFF2CC",
        "Applied": "C6E0B4",
        "Rejected": "F4B084",
        "Interview": "9DC3E6",
    }
    h1b_yes_fill = PatternFill("solid", fgColor="C6E0B4")
    link_font = Font(color="0563C1", underline="single")

    link_col = COLUMNS.index("Apply Link") + 1
    status_col = COLUMNS.index("Status") + 1
    h1b_col = COLUMNS.index("H1B Sponsor?") + 1

    for i, r in enumerate(rows, start=2):
        ws.append([r.get(c, "") for c in COLUMNS])

        url = r.get("Apply Link") or ""
        if url.startswith(("http://", "https://")):
            cell = ws.cell(row=i, column=link_col)
            cell.hyperlink = url
            cell.font = link_font

        status = r.get("Status") or "New"
        if status in status_colors:
            ws.cell(row=i, column=status_col).fill = PatternFill(
                "solid", fgColor=status_colors[status]
            )

        if (r.get("H1B Sponsor?") or "").lower() == "yes":
            ws.cell(row=i, column=h1b_col).fill = h1b_yes_fill

    if rows:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_VALUES)}"',
            allow_blank=False,
        )
        dv.add(f"H2:H{len(rows) + 1}")
        ws.add_data_validation(dv)

    widths = {
        "A": 38, "B": 26, "C": 22, "D": 13, "E": 55,
        "F": 7, "G": 12, "H": 12, "I": 13, "J": 30, "K": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.save(TRACKER_XLSX)


def main() -> int:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("dashboard.json")
    if not src.exists():
        print(f"ERROR: {src} not found", file=sys.stderr)
        return 1

    dashboard = json.loads(src.read_text(encoding="utf-8"))
    fresh = dashboard.get("top_jobs") or []
    today = datetime.utcnow().strftime("%Y-%m-%d")

    sponsors = load_h1b_sponsors()
    existing = load_existing_tracker()

    by_key: dict[tuple[str, str], dict] = dict(existing)

    # Merge fresh leads
    for job in fresh:
        title = (job.get("title") or "").strip()
        company = (job.get("company") or "").strip()
        if not title or not company:
            continue
        key = (title.lower(), company.lower())
        if key in by_key:
            row = by_key[key]
            # Refresh dynamic fields, preserve user-edited fields
            new_url = job.get("url")
            if new_url:
                row["Apply Link"] = new_url
            row["Score"] = str(job.get("fit_score") or row.get("Score") or "")
            row["Confidence"] = job.get("confidence") or row.get("Confidence") or ""
            if not row.get("First Seen"):
                row["First Seen"] = today
            if not row.get("Status"):
                row["Status"] = "New"
            row["H1B Sponsor?"] = "Yes" if is_h1b_sponsor(company, sponsors) else "No"
        else:
            by_key[key] = {
                "Title": title,
                "Company": company,
                "Location": job.get("location") or "",
                "H1B Sponsor?": "Yes" if is_h1b_sponsor(company, sponsors) else "No",
                "Apply Link": job.get("url") or "",
                "Score": str(job.get("fit_score") or ""),
                "Confidence": job.get("confidence") or "",
                "Status": "New",
                "Applied Date": "",
                "Notes": "",
                "First Seen": today,
            }

    # Sort: New first (highest score), then Interview, Applied, Rejected
    def sort_key(r: dict) -> tuple[int, float]:
        try:
            score = -float(r.get("Score") or 0)
        except (TypeError, ValueError):
            score = 0.0
        status_idx = STATUS_ORDER.get((r.get("Status") or "New"), 0)
        return (status_idx, score)

    rows = sorted(by_key.values(), key=sort_key)

    write_csv(rows)
    write_xlsx(rows)

    counts = {s: 0 for s in STATUS_VALUES}
    for r in rows:
        counts[r.get("Status") or "New"] = counts.get(r.get("Status") or "New", 0) + 1
    print(
        f"Tracker updated: {len(rows)} total | "
        f"New={counts['New']} Applied={counts['Applied']} "
        f"Interview={counts['Interview']} Rejected={counts['Rejected']}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
