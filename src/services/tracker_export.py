"""Application tracker export service.

Exports job leads, contacts, and application status to CSV
or Excel spreadsheets for offline tracking.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional


def export_tracker_csv(
    leads: list[dict],
    contacts: list[dict],
    output_path: str = "application_tracker.csv",
) -> str:
    """Export an application tracker to CSV.

    Columns: Company | Role | Fit Score | Priority | Apply URL |
             Contact Person | Contact Role | Outreach Type | Status | Notes

    Returns the output file path.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # Build a contact lookup by company
    contact_by_company: dict[str, list[dict]] = {}
    for c in contacts:
        company = c.get("company", "")
        if company not in contact_by_company:
            contact_by_company[company] = []
        contact_by_company[company].append(c)

    rows = []
    for lead in leads:
        company = lead.get("company", "")
        score = lead.get("fit_score", 0)
        priority = _score_to_priority(score)

        # Find best contact for this company
        company_contacts = contact_by_company.get(company, [])
        if company_contacts:
            # Pick highest-priority contact
            best = sorted(
                company_contacts,
                key=lambda c: {"high": 0, "medium": 1, "low": 2}.get(
                    c.get("contact_priority", "medium"), 1
                ),
            )[0]
            contact_name = best.get("name", "")
            contact_role = best.get("role", "")
            outreach_type = (best.get("suggested_outreach_type") or "").replace("_", " ")
        else:
            contact_name = ""
            contact_role = ""
            outreach_type = ""

        rows.append({
            "Company": company,
            "Role": lead.get("title", ""),
            "Fit Score": f"{score:.0f}",
            "Priority": priority,
            "Apply URL": lead.get("url", ""),
            "Contact Person": contact_name,
            "Contact Role": contact_role,
            "Outreach Type": outreach_type,
            "Status": "Not Applied",
            "Notes": "; ".join(lead.get("match_reasons", [])[:2]),
        })

    # Sort by fit score descending
    rows.sort(key=lambda r: float(r["Fit Score"]), reverse=True)

    fieldnames = [
        "Company", "Role", "Fit Score", "Priority", "Apply URL",
        "Contact Person", "Contact Role", "Outreach Type", "Status", "Notes",
    ]

    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return str(out)


def export_tracker_excel(
    leads: list[dict],
    contacts: list[dict],
    output_path: str = "application_tracker.xlsx",
) -> str:
    """Export an application tracker to Excel with formatting.

    Falls back to CSV if openpyxl is not installed.
    Returns the output file path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        # Fall back to CSV
        csv_path = output_path.replace(".xlsx", ".csv")
        return export_tracker_csv(leads, contacts, csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Application Tracker"

    # Headers
    headers = [
        "Company", "Role", "Fit Score", "Priority", "Apply URL",
        "Contact Person", "Contact Role", "Outreach Type", "Status", "Notes",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Build contact lookup
    contact_by_company: dict[str, list[dict]] = {}
    for c in contacts:
        company = c.get("company", "")
        if company not in contact_by_company:
            contact_by_company[company] = []
        contact_by_company[company].append(c)

    # Sort leads by score
    sorted_leads = sorted(leads, key=lambda l: l.get("fit_score", 0), reverse=True)

    # Priority fill colors
    priority_fills = {
        "HIGH": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "LOW": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row_idx, lead in enumerate(sorted_leads, 2):
        company = lead.get("company", "")
        score = lead.get("fit_score", 0)
        priority = _score_to_priority(score)

        company_contacts = contact_by_company.get(company, [])
        if company_contacts:
            best = sorted(
                company_contacts,
                key=lambda c: {"high": 0, "medium": 1, "low": 2}.get(
                    c.get("contact_priority", "medium"), 1
                ),
            )[0]
            contact_name = best.get("name", "")
            contact_role = best.get("role", "")
            outreach_type = (best.get("suggested_outreach_type") or "").replace("_", " ")
        else:
            contact_name = ""
            contact_role = ""
            outreach_type = ""

        row_data = [
            company,
            lead.get("title", ""),
            score,
            priority,
            lead.get("url", ""),
            contact_name,
            contact_role,
            outreach_type,
            "Not Applied",
            "; ".join(lead.get("match_reasons", [])[:2]),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col == 4:  # Priority column
                fill = priority_fills.get(priority)
                if fill:
                    cell.fill = fill

    # Auto-fit column widths (approximate)
    col_widths = [25, 35, 10, 10, 45, 25, 30, 18, 15, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def _score_to_priority(score: float) -> str:
    """Convert fit score to priority label."""
    if score >= 85:
        return "HIGH"
    elif score >= 60:
        return "MEDIUM"
    return "LOW"
